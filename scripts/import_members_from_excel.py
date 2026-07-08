from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app import scrim_db


COLUMN_MAP = {
    "name": "이름",
    "participation_status_text": "참가여부",
    "absence_reason": "불참사유",
    "payment_status": "입금",
    "riot_id": "아이디",
    "tier": "티어",
    "top_adjustment": "탑레조정",
    "game_count_adjustment": "판수조정",
    "preferred_lines": "참가라인",
    "score_top": "탑",
    "score_jungle": "정글",
    "score_mid": "미드",
    "score_adc": "원딜",
    "score_support": "서폿",
    "notes": "기타",
}


def clean(value) -> str:
    return str(value or "").strip()


def split_riot_ids(raw: str) -> tuple[str | None, str | None]:
    parts = [part.strip() for part in raw.replace("\n", ",").split(",") if part.strip()]
    if not parts:
        return None, None
    return parts[0], parts[1] if len(parts) > 1 else None


def worksheet_for_roster(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    return workbook["목록"] if "목록" in workbook.sheetnames else workbook.active


def rows_from_workbook(path: Path) -> list[dict]:
    sheet = worksheet_for_roster(path)
    headers = [clean(cell.value) for cell in next(sheet.iter_rows(max_row=1))]
    index_by_header = {header: index for index, header in enumerate(headers) if header}
    missing = [label for label in ("이름", "아이디") if label not in index_by_header]
    if missing:
        raise ValueError(f"필수 컬럼이 없습니다: {', '.join(missing)}")

    rows = []
    for source_row, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {}
        for field, header in COLUMN_MAP.items():
            index = index_by_header.get(header)
            values[field] = clean(row[index].value) if index is not None else None
        if not values["name"]:
            continue
        riot_id, secondary_riot_id = split_riot_ids(values.get("riot_id") or "")
        values["riot_id"] = riot_id
        values["secondary_riot_id"] = secondary_riot_id
        values["source_row"] = source_row
        rows.append(values)
    return rows


def import_rows(rows: list[dict], *, dry_run: bool) -> dict:
    summary = {
        "total_rows": len(rows),
        "with_riot_id": sum(1 for row in rows if row.get("riot_id")),
        "without_riot_id": sum(1 for row in rows if not row.get("riot_id")),
        "created_or_updated": 0,
        "accounts_issued": 0,
    }
    if dry_run:
        return summary

    scrim_db.init_db()
    with scrim_db.connect() as connection:
        for row in rows:
            entry = scrim_db.upsert_roster_entry(connection, **row)
            summary["created_or_updated"] += 1
            if entry.get("account_status") == "ISSUED":
                summary["accounts_issued"] += 1
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Import the 수강대난투 roster sheet into the local or configured scrim DB.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows = rows_from_workbook(args.workbook)
    summary = import_rows(rows, dry_run=args.dry_run)
    for key, value in summary.items():
        print(f"{key}={value}")


if __name__ == "__main__":
    main()
